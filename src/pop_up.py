import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
import re
from PyQt6.QtWidgets import (
    QVBoxLayout, QLabel, QApplication,
    QPushButton, QHBoxLayout, QFileDialog, QMessageBox, QProgressBar, QDialog, QLineEdit
)
from PyQt6.QtCore import Qt
import platform
import subprocess
import time

from pdf import (extract_text_from_pdf, summarize_text)
from job_matcher import check_predicted_job_similarity


class BatchProcessingDialog(QDialog):
    def __init__(self, embeddings, jobs_excel):
        """
        Initialize the BatchProcessingDialog.

        This dialog allows the user to select a folder containing PDFs
        to process, specify an output file name, and start the processing
        of the PDFs.

        Parameters:
        - embeddings (list): precomputed embeddings for the job offers
        - jobs_excel (str): path to the Excel file containing job offers
        """
        super().__init__()
        self.setWindowTitle("Process Multiple PDFs")
        self.setGeometry(100, 100, 400, 250)
        self.embeddings = embeddings
        self.jobs_excel = jobs_excel

        self.layout = QVBoxLayout(self)

        self.folder_input = os.getenv("PATH_FOLDER_PDFS")

        # Folder input
        self.folder_label = QLabel(f"Selected Folder: {self.folder_input}", self)
        self.layout.addWidget(self.folder_label)

        self.select_folder_button = QPushButton("Select Folder", self)
        self.select_folder_button.clicked.connect(self.select_folder)
        self.layout.addWidget(self.select_folder_button)

        self.output_directory = self.folder_input

        # Output file input
        self.file_label = QLabel("Enter Output File Name:", self)
        self.layout.addWidget(self.file_label)

        output_layout = QHBoxLayout()

        self.file_input = QLineEdit(os.getenv("FILE_EXCEL_RESULTS"), self)
        output_layout.addWidget(self.file_input)

        self.output_dir_button = QPushButton("Select Output Directory", self)
        self.output_dir_button.clicked.connect(self.select_output_directory)
        output_layout.addWidget(self.output_dir_button)

        self.layout.addLayout(output_layout)
        self.output_path_label = QLabel(f"Output Directory: {self.output_directory}", self)
        self.layout.addWidget(self.output_path_label)

        # Progress bar
        self.progress_bar = QProgressBar(self)
        self.progress_bar.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.layout.addWidget(self.progress_bar)

        self.remaining_time_label = QLabel("")
        self.layout.addWidget(self.remaining_time_label)

        # Start button
        self.start_button = QPushButton("Start Processing", self)
        self.start_button.clicked.connect(self.start_processing)
        self.layout.addWidget(self.start_button)

    def select_input_folder(self):
        """
        Opens a file dialog for the user to select a folder
        containing PDF files to process.

        If a folder is selected, the text field for the input folder
        is updated with the selected folder path.
        """
        folder = QFileDialog.getExistingDirectory(self, "Select Input Folder")
        if folder:
            self.folder_input.setText(folder)

    def start_processing(self):
        """
        Starts the batch processing of PDF files in the selected input folder.

        This function:

        1. Checks if the input folder is valid and if the output file already exists.
        2. Opens the input folder and iterates over the PDF files.
        3. Extracts the text from each PDF file, summarizes it, and checks the similarity with the job offers.
        4. Saves the results in the specified output Excel file.
        5. Reports the summary of the processing.

        :return: None
        """
        self.remaining_time_label.setText("Estimated time remaining: calculating...")

        input_directory = self.folder_input.strip()
        output_excel = os.path.join(self.output_directory, self.file_input.text())

        if not os.path.exists(input_directory):
            QMessageBox.critical(self, "Error", "Invalid input folder selected.")
            return

        if os.path.exists(output_excel):
            os.remove(output_excel)

        wb = Workbook()
        ws = wb.active
        ws.title = "Job Matches"
        ws.append(["File Name","Similarity", "Details"])

        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        orange_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
        green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

        def extract_similarity(result_text):
            """
            Extract the similarity percentage from a job matching result text.

            Given a text containing a job matching result, extract the similarity percentage from it.
            The percentage is either extracted from a substring like "XX%" or from a substring
            like "similarity near zero percent". If no percentage is found, None is returned.

            Parameters:
            - result_text (str): The text containing the job matching result.

            Returns:
            - int or None: The extracted similarity percentage or None if no percentage was found.
            """
            match = re.search(r"(\d+)%", result_text)
            if match:
                return int(match.group(1))
            if re.search(r"similarity.*?near zero percent", result_text, re.IGNORECASE):
                return 0
            return None

        def sort_key(file_name):
            """
            Generate a sorting key for a given file name.

            This function parses the file name to extract an alphabetical prefix and a numerical suffix.
            The prefix is converted to lowercase for consistent sorting, and the numerical suffix is
            converted to an integer for numerical sorting. If no numerical suffix is present, infinity is
            used to ensure the prefix is sorted before any numbered variants.

            Parameters:
            - file_name (str): The name of the file to generate a sorting key for.

            Returns:
            - tuple: A tuple consisting of the lowercase prefix and the numerical suffix, or infinity if
              no numerical suffix is found.
            """
            match = re.match(r'([a-zA-Z_]*)(\d*)', file_name)
            if match:
                prefix = match.group(1)
                number = match.group(2)
                number = int(number) if number else float('inf')
                return prefix.lower(), number
            else:
                return file_name.lower(), float('inf')

        files = sorted(
            [f for f in os.listdir(input_directory) if f.endswith(".pdf")],
            key=sort_key
        )

        self.progress_bar.setMaximum(len(files))

        low_similarity = 0
        medium_similarity = 0
        high_similarity = 0

        start_time = time.time()

        for i, file_name in enumerate(files):
            pdf_path = os.path.join(input_directory, file_name)
            try:
                extracted_text = extract_text_from_pdf(pdf_path)
                cv_text = summarize_text(extracted_text)
                similarity_result = check_predicted_job_similarity(cv_text, self.jobs_excel, self.embeddings)
                similarity_percentage = extract_similarity(similarity_result)

                row = [file_name, f"{similarity_percentage}%", similarity_result]
                ws.append(row)

                file_cell = ws.cell(row=ws.max_row, column=1)
                file_cell.alignment = Alignment(horizontal="center", vertical="center")

                file_cell = ws.cell(row=ws.max_row, column=2)
                file_cell.alignment = Alignment(horizontal="center", vertical="center")

                if similarity_percentage is not None:
                    color = None
                    if similarity_percentage < 50:
                        color = red_fill
                        low_similarity += 1
                    elif 50 <= similarity_percentage < 60:
                        color = orange_fill
                        medium_similarity += 1
                    elif similarity_percentage >= 60:
                        color = green_fill
                        high_similarity += 1

                    ws[f"A{ws.max_row}"].fill = color

                details_cell = ws.cell(row=ws.max_row, column=3)
                details_cell.alignment = Alignment(wrap_text=True)
            except Exception as e:
                print(f"Error processing {file_name}: {e}")

            elapsed_time = time.time() - start_time
            files_processed = i + 1
            avg_time_per_file = elapsed_time / files_processed
            remaining_files = len(files) - files_processed
            estimated_remaining_time = avg_time_per_file * remaining_files

            if estimated_remaining_time >= 60:
                remaining_minutes = int(estimated_remaining_time // 60)
                remaining_seconds = int(estimated_remaining_time % 60)
                self.remaining_time_label.setText(
                    f"Estimated time remaining: {remaining_minutes} minutes {remaining_seconds} seconds"
                )
            else:
                self.remaining_time_label.setText(
                    f"Estimated time remaining: {int(estimated_remaining_time)} seconds"
                )

            self.progress_bar.setValue(i + 1)

            QApplication.processEvents()

        elapsed_time = time.time() - start_time
        elapsed_minutes = int(elapsed_time // 60)
        elapsed_seconds = int(elapsed_time % 60)
        if elapsed_minutes == 0:
            self.remaining_time_label.setText(
                f"Processing complete! Time taken: {elapsed_seconds} seconds."
            )
        else:
            self.remaining_time_label.setText(
                f"Processing complete! Time taken: {elapsed_minutes} minutes {elapsed_seconds} seconds."
            )

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2 # max(min(max_length + 2, 50), 10)

        wb.save(output_excel)

        try:
            if platform.system() == "Darwin":  # macOS
                subprocess.run(["open", output_excel], check=True)
            elif platform.system() == "Windows":  # Windows
                os.startfile(output_excel)
            elif platform.system() == "Linux":  # Linux
                subprocess.run(["xdg-open", output_excel], check=True)
            else:
                QMessageBox.warning(self, "Warning", "Could not determine your operating system to open the file.")
        except Exception as e:
            QMessageBox.warning(self, "Warning", f"Unable to open the file automatically: {e}")

        # Report Summary
        total_files = low_similarity + medium_similarity + high_similarity
        report = (
            f"--- Processing Report ---\n"
            f"Total files processed: {total_files}\n"
            f"Low similarity (< 50%): {low_similarity}\n"
            f"Medium similarity (50%-60%): {medium_similarity}\n"
            f"High similarity (>= 60%): {high_similarity}\n"
            f"\nResults saved to\n{output_excel}"
        )
        QMessageBox.information(self, "Success", report)

    def select_folder(self):
        """
        Opens a file dialog to select a folder containing PDF files to process.

        If a folder is selected, the text field for the input folder is updated with the selected folder path.
        Additionally, the output directory is updated to the selected folder path.
        """
        folder = QFileDialog.getExistingDirectory(self, "Select Input Folder", self.folder_input)
        if folder:
            self.folder_label.setText(f"Selected Folder: {folder}")
            self.folder_input = folder
            self.output_directory = folder
            self.output_path_label.setText(f"Output Directory: {self.output_directory}")

    def select_output_directory(self):
        """
        Opens a file dialog to select a directory where the output Excel file will be saved.

        If a directory is selected, the text field for the output directory is updated with the selected directory path.

        :return: None
        """
        directory = QFileDialog.getExistingDirectory(self, "Select Output Directory", self.output_directory)
        if directory:
            self.output_directory = directory
            self.output_path_label.setText(f"Output Directory: {self.output_directory}")


